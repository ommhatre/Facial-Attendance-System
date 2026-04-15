from flask import Flask, render_template, Response, request
import cv2
import os
import face_recognition
from datetime import datetime
import openpyxl
from deepface import DeepFace
from collections import defaultdict
from ast import literal_eval
import statistics
import glob

app = Flask(__name__)

# Load user face encodings from 'users' folder
user_face_encodings = []
user_face_names = []
user_data = 'users'
for u in os.listdir(user_data):
    path = os.path.join(user_data, u)
    name = os.path.splitext(u)[0]
    user_image = face_recognition.load_image_file(path)
    user_encoding = face_recognition.face_encodings(user_image)[0]
    user_face_encodings.append(user_encoding)
    user_face_names.append(name)

# Excel workbook (used only when marking)
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Name", "Time", "Dominant Emotion", "All Emotions", "Productivity (%)"])

def is_name_logged(name):
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == name:
            return True
    return False

def calculate_productivity_percentage(emotions):
    weights = {
        'happy': 2,
        'neutral': 1,
        'surprise': 0,
        'sad': -1,
        'angry': -2,
        'fear': -2,
        'disgust': -2
    }
    score = sum(weights.get(e, 0) * v for e, v in emotions.items())
    total_weight = sum(emotions.values())
    if total_weight == 0:
        return 50.0
    normalized_score = score / (2 * total_weight)
    return round((normalized_score + 1) / 2 * 100, 2)

def detect_faces():
    camera = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    while True:
        ret, frame = camera.read()
        if not ret:
            break
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(frame_rgb)
        face_encodings = face_recognition.face_encodings(frame_rgb, face_locations)
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            matches = face_recognition.compare_faces(user_face_encodings, face_encoding)
            name = "Unknown"
            color = (0, 0, 255)  # Default to red for unknown

            if True in matches:
                first_match_index = matches.index(True)
                name = user_face_names[first_match_index]
                color = (0, 255, 0)  # Green for recognized
                if not is_name_logged(name):
                    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    face_crop = frame[top:bottom, left:right]
                    try:
                        analysis = DeepFace.analyze(face_crop, actions=["emotion"], enforce_detection=False)
                        dominant_emotion = analysis[0]['dominant_emotion']
                        all_emotions = analysis[0]['emotion']
                    except:
                        dominant_emotion = "Error"
                        all_emotions = {}
                    productivity = calculate_productivity_percentage(all_emotions)
                    ws.append([name, current_time, dominant_emotion, str(all_emotions), productivity])
                    log_name = f"log_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
                    wb.save(log_name)

            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
        
        ret, jpeg = cv2.imencode('.jpg', frame)
        frame_bytes = jpeg.tobytes()
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    camera.release()


@app.route('/')
def home():
    return render_template('index.html')

@app.route('/mark_attendance')
def mark_attendance():
    return render_template('mark_attendance.html')

@app.route('/video_feed')
def video_feed():
    return Response(detect_faces(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/display_attendance', methods=['GET', 'POST'])
def display_attendance():
    available_logs = sorted([f.split("_")[1].split(".")[0] for f in glob.glob("log_*.xlsx")])
    selected_date = request.form.get("date") if request.method == "POST" else available_logs[-1] if available_logs else None
    data = []
    if selected_date:
        file_path = f"log_{selected_date}.xlsx"
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=5):
                name, time, emotion, emotions_dict, productivity = [cell.value for cell in row]
                data.append((name, time, emotion, emotions_dict, productivity))
    return render_template('display_attendance.html', data=data, available_logs=available_logs, selected_date=selected_date)

@app.route('/productivity')
def productivity():
    user_emotions = defaultdict(lambda: defaultdict(float))  # user -> emotion -> total
    user_daily_productivity = defaultdict(lambda: defaultdict(float))  # user -> date -> avg_prod

    for filename in glob.glob("log_*.xlsx"):
        date_str = filename.split("_")[1].split(".")[0]
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        user_day_sum = defaultdict(list)

        for row in ws.iter_rows(min_row=2, max_col=5):
            name, _, _, emotions_raw, productivity = [cell.value for cell in row]
            try:
                emotions = literal_eval(emotions_raw)
            except:
                continue

            for emotion, val in emotions.items():
                user_emotions[name][emotion] += val

            try:
                prod = float(productivity)
                user_day_sum[name].append(prod)
            except:
                continue

        for name, prods in user_day_sum.items():
            if prods:
                avg = round(sum(prods) / len(prods), 2)
                user_daily_productivity[name][date_str] = avg

    user_data = {}
    for user in user_emotions:
        user_data[user] = {
            "emotions": dict(user_emotions[user]),
            "daily_productivity": dict(user_daily_productivity[user])
        }

    return render_template("productivity.html", user_data=user_data)


if __name__ == '__main__':
    app.run(debug=True)
